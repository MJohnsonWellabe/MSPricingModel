"""The assumptions Excel export builds a well-formed workbook whose raw values and
derived factors match the source AssumptionSet."""
from __future__ import annotations

import io

import openpyxl
import pytest

from medigap_engine.io.excel_export import (
    SHEET_NAMES,
    assumptions_to_xlsx_bytes,
)
from medigap_engine.models.assumptions import derive_two_level, normalized_factors


def _load(asm):
    return openpyxl.load_workbook(io.BytesIO(assumptions_to_xlsx_bytes(asm)))


def _find_value(ws, label):
    """Value in column B on the row whose column A equals ``label``."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return ws.cell(row=r, column=2).value
    return None


def test_workbook_has_all_sheets(asm):
    wb = _load(asm)
    assert wb.sheetnames == SHEET_NAMES


def test_raw_values_present(asm):
    wb = _load(asm)
    assert _find_value(wb["Pull forward"], "Duration (years to pull forward)") == asm.pull_forward.duration
    # a base-premium row (issue age -> base) lives on the Premium sheet
    ws = wb["Premium"]
    issue_age, base = sorted(asm.premium.base_by_issue_age.items())[0]
    found = False
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == issue_age and ws.cell(row=r, column=2).value == base:
            found = True
            break
    assert found, "expected base premium row not found on Premium sheet"


def test_derived_bring_forward_factor(asm):
    wb = _load(asm)
    pf = asm.pull_forward
    expected = (1.0 + pf.claims_trend) ** pf.duration
    assert _find_value(wb["Derived factors"], "Claims bring-forward") == pytest.approx(expected)


def test_derived_premium_gender_factors_normalised(asm):
    wb = _load(asm)
    ws = wb["Derived factors"]
    expected = normalized_factors(
        {"M": 1.0 + asm.premium.gender_diff, "F": 1.0}, asm.distribution.gender)
    got = {}
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Gender" and ws.cell(row=r, column=2).value in ("M", "F"):
            got[ws.cell(row=r, column=2).value] = ws.cell(row=r, column=3).value
        if set(got) == {"M", "F"}:
            break
    assert got["M"] == pytest.approx(expected["M"])
    assert got["F"] == pytest.approx(expected["F"])


def test_derived_morbidity_two_level_factors(asm):
    wb = _load(asm)
    ws = wb["Derived factors"]
    pref = derive_two_level(asm.distribution.preferred.get("Y", 0.5), asm.morbidity.preferred_diff)
    # the morbidity Preferred Y/N factors should appear on the sheet
    seen = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=1).value == "Preferred"]
    assert any(v == pytest.approx(pref["Y"]) for v in seen)
    assert any(v == pytest.approx(pref["N"]) for v in seen)


def test_distribution_sums_written(asm):
    wb = _load(asm)
    assert _find_value(wb["Distribution"], "Sum") == pytest.approx(
        sum(asm.distribution.by_issue_age.values()))
