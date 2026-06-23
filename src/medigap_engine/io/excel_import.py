"""Parse an assumptions workbook produced by :mod:`excel_export` back into the
canonical assumptions dict (the same shape as :func:`serialize.assumptions_to_dict`),
so callers can rebuild an :class:`AssumptionSet` via ``assumptions_from_dict``.

The parser mirrors the ``_kv`` / ``_table`` layout of ``excel_export``: each sheet
is a sequence of titled blocks (a bold title in column A, then either label/value
rows or a header row followed by data rows, ending at a blank row). It matches
blocks by their known titles/labels. The output-only **Derived factors** sheet is
ignored. ``openpyxl`` is imported lazily so the engine still imports numpy-only.
"""
from __future__ import annotations

import re

from .excel_export import SHEET_NAMES

_PLAN_GRID_RE = re.compile(r"^Plan (.+) weight grid \(issue age x UW\)$")


# --- row-scanning helpers (operate on a list of value-tuples for one sheet) ---

def _kv(rows, label):
    """Value in column B on the row whose column A equals ``label`` (or None)."""
    for r in rows:
        if r and r[0] == label:
            return r[1] if len(r) > 1 else None
    return None


def _table(rows, title):
    """(header tuple, list of data rows) for the titled block, or (None, [])."""
    for i, r in enumerate(rows):
        if r and r[0] == title:
            header = rows[i + 1] if i + 1 < len(rows) else ()
            data = []
            for dr in rows[i + 2:]:
                if not dr or dr[0] is None:
                    break
                data.append(dr)
            return header, data
    return None, []


def _ncols(header) -> int:
    n = 0
    for c in header or ():
        if c is None:
            break
        n += 1
    return n


def _trim_trailing_none(seq) -> list:
    out = list(seq)
    while out and out[-1] is None:
        out.pop()
    return out


def _factor_dict(data) -> dict:
    """{col-A -> float(col-B)} skipping rows with a blank key or value (the seed
    carries a couple of non-numeric artefact keys that the JSON path tolerates)."""
    out = {}
    for r in data:
        if r[0] is None or r[0] == "Sum" or r[1] is None:
            continue
        out[str(r[0])] = float(r[1])
    return out


# --- per-sheet parsers ------------------------------------------------------

def _pull_forward(rows) -> dict:
    return {
        "duration": float(_kv(rows, "Duration (years to pull forward)")),
        "claims_trend": float(_kv(rows, "Claims trend (pull-forward)")),
        "premium_trend": float(_kv(rows, "Premium trend (pull-forward)")),
    }


def _morbidity(rows) -> dict:
    header, data = _table(rows, "Base claim cost by attained age and plan (gender blend)")
    plans = [str(c) for c in header[1:_ncols(header)]]
    ages = [int(r[0]) for r in data]
    base_cc = {pl: [float(r[1 + j]) for r in data] for j, pl in enumerate(plans)}

    _h, td = _table(rows, "Projection trend & claim-cost aging by duration")
    trend_by_year = [float(r[1]) for r in td]
    cc_aging = [float(r[2]) for r in td]

    _h, sd = _table(rows, "State morbidity factors")
    state_factors = _factor_dict(sd)

    _h, sel = _table(rows, "UW selection factors")
    selection_factors = [
        {"duration": int(r[0]), "issue_age": int(r[1]), "uw": str(r[2]), "factor": float(r[3])}
        for r in sel]

    _h, rf = _table(rows, "Raw preferred / HHD claim factors by level")
    preferred_factors = {str(r[0]): float(r[1]) for r in rf if r[1] is not None}
    hhd_factors = {str(r[0]): float(r[2]) for r in rf if len(r) > 2 and r[2] is not None}

    return {
        "ages": ages, "plans": plans, "base_cc": base_cc,
        "gender_cc_diff": float(_kv(rows, "Male claim cost higher than female by")),
        "preferred_diff": float(_kv(rows, "Non-preferred higher than preferred by")),
        "hhd_diff": float(_kv(rows, "Non-HHD higher than HHD by")),
        "state_factors": state_factors,
        "selection_factors": selection_factors,
        "cc_aging_by_duration": cc_aging,
        "trend_by_year": trend_by_year,
        "preferred_factors": preferred_factors,
        "hhd_factors": hhd_factors,
    }


def _premium(rows) -> dict:
    _h, base = _table(rows, "Base premium by issue age (plan-G blend)")
    _h, plan = _table(rows, "Plan relativities (G = 1.00)")
    _h, uw = _table(rows, "UW relativities")
    _h, state = _table(rows, "State premium factors (raw)")
    cp_header, cp = _table(
        rows, "Per-cell premiums — exact rates by state (override the factor model)")
    cell_premiums: dict[str, dict[str, float]] = {}
    cp_states = [str(c) for c in (cp_header or ())[1:_ncols(cp_header)]]
    for r in cp:
        label = str(r[0])
        cell_premiums[label] = {s: float(r[1 + j]) for j, s in enumerate(cp_states)
                                if r[1 + j] is not None}
    return {
        "base_by_issue_age": {int(r[0]): float(r[1]) for r in base},
        "plan_rel": {str(r[0]): float(r[1]) for r in plan},
        "uw_rel": {str(r[0]): float(r[1]) for r in uw},
        "gender_diff": float(_kv(rows, "Male premium higher than female by")),
        "preferred_diff": float(_kv(rows, "Non-preferred premium higher by")),
        "hhd_diff": float(_kv(rows, "Non-HHD premium higher by")),
        "state_factor": _factor_dict(state),
        "cell_premiums": cell_premiums,
    }


def _rerates(rows) -> dict:
    _h, spec = _table(rows, "Specified rerates by duration (shared / national)")
    _h, aging = _table(rows, "Aging rerate by attained age (column H)")
    irr = _kv(rows, "Target IRR (reported)")
    bs_header, bs = _table(rows, "Per-state rerate overrides by duration")
    bs_states = [str(c) for c in (bs_header or ())[1:_ncols(bs_header)]]
    by_state = {s: [float(r[1 + j]) for r in bs] for j, s in enumerate(bs_states)
                if all(r[1 + j] is not None for r in bs)}
    _h, tgt = _table(rows, "Per-state target lifetime LR overrides")
    target_by_state = {str(r[0]): float(r[1]) for r in tgt if r[0] is not None}
    return {
        "solve": bool(_kv(rows, "Solve to target lifetime LR")),
        "specified_rerates": [float(r[1]) for r in spec],
        "aging_rerate_by_age_ages": [int(r[0]) for r in aging],
        "aging_rerate_by_age_factor": [float(r[1]) for r in aging],
        "target_lifetime_lr": float(_kv(rows, "Target lifetime LR")),
        "target_irr": (None if irr is None else float(irr)),
        "max_rerate": float(_kv(rows, "Max single rerate")),
        "in_year_lr_floor": float(_kv(rows, "In-year LR floor")),
        "consecutive_z": float(_kv(rows, "Consecutive rule: z")),
        "consecutive_b": int(_kv(rows, "Consecutive rule: b (years)")),
        "antiselection_lambda_claims": float(_kv(rows, "Antiselection lambda — claims")),
        "antiselection_lambda_lapse": float(_kv(rows, "Antiselection lambda — lapse")),
        "by_state": by_state,
        "target_lifetime_lr_by_state": target_by_state,
    }


def _marginal(rows, title) -> dict:
    _h, data = _table(rows, title)
    return _factor_dict(data)


def _distribution(rows) -> dict:
    joint: dict[str, dict[str, dict[str, float]]] = {}
    for i, r in enumerate(rows):
        if not r or not isinstance(r[0], str):
            continue
        m = _PLAN_GRID_RE.match(r[0])
        if not m:
            continue
        pl = m.group(1)
        header, data = _table(rows, r[0])
        uws = [str(c) for c in header[1:_ncols(header)]]
        grid: dict[str, dict[str, float]] = {}
        for dr in data:
            age = str(int(dr[0]))
            grid[age] = {u: float(dr[1 + j]) for j, u in enumerate(uws)}
        joint[pl] = grid
    # per-state mix-of-business: joint grid (state, plan, issue age, *uw) ...
    bsj_header, bsj = _table(rows,
                             "Per-state mix-of-business grid (joint plan x issue age x UW)")
    bs_uws = [str(c) for c in (bsj_header or ())[3:_ncols(bsj_header)]]
    by_state: dict[str, dict] = {}
    for r in bsj:
        s, pl, age = str(r[0]), str(r[1]), str(int(r[2]))
        cell = {u: float(r[3 + j]) for j, u in enumerate(bs_uws)}
        by_state.setdefault(s, {}).setdefault("joint", {}).setdefault(pl, {})[age] = cell
    # ... and the per-state gender / preferred / HHD marginals layered on top
    _h, marg = _table(rows, "Per-state gender / preferred / HHD marginals")
    for r in marg:
        s, dim, lvl, w = str(r[0]), str(r[1]), str(r[2]), float(r[3])
        by_state.setdefault(s, {}).setdefault(dim, {})[lvl] = w

    _h, weights = _table(rows, "New-business volume weights by state (combine weighting)")
    state_weights = {str(r[0]): float(r[1]) for r in weights if r[0] is not None}
    _h, sep = _table(rows, "Special Enrollment Period (SEP) states")
    sep_states = [str(r[0]) for r in sep if r[0] is not None]
    return {
        "joint": joint,
        "gender": _marginal(rows, "Gender"),
        "preferred": _marginal(rows, "Preferred"),
        "hhd": _marginal(rows, "HHD"),
        "by_state": by_state,
        "state_weights": state_weights,
        "sep_rule_states": sep_states,
    }


def _termination(rows) -> dict:
    _h, lt = _table(rows, "Base lapse (blend) and UW relativity by duration")
    _h, st = _table(rows, "State lapse factors")
    _h, mort = _table(rows, "Mortality table")
    return {
        "base_lapse": [float(r[1]) for r in lt],
        "uw_lapse_rel": [float(r[2]) for r in lt],
        "state_factors": _factor_dict(st),
        "mort_age": [int(r[0]) for r in mort],
        "mort_qx": [float(r[1]) for r in mort],
        "dur2_scaling": float(_kv(rows, "Duration 2 scaling")),
        "dur3plus_scaling": float(_kv(rows, "Duration 3+ scaling")),
    }


def _commission(rows) -> dict:
    header, data = _table(rows, "Commission rate by state and duration")
    states = [str(c) for c in header[1:_ncols(header)]]
    by_state = {}
    for j, s in enumerate(states):
        by_state[s] = _trim_trailing_none([float(r[1 + j]) if r[1 + j] is not None else None
                                           for r in data])
    _h, nat = _table(rows, "National schedules by duration")
    plan_n = _trim_trailing_none([r[1] for r in nat])
    nonn = _trim_trailing_none([r[2] for r in nat])
    return {
        "by_state": {s: [float(v) for v in vs] for s, vs in by_state.items()},
        "plan_n_schedule": [float(v) for v in plan_n],
        "nonn_schedule": [float(v) for v in nonn],
        "gi_flat": float(_kv(rows, "GI flat commission")),
        "plan_f_offset": float(_kv(rows, "Plan F premium offset")),
        "age80_halving": bool(_kv(rows, "Halve commission for issue age >= 80")),
    }


_ECONOMIC_LABELS = {
    "discount_rate": "Discount rate",
    "nier": "NIER (investment return)",
    "inflation": "Inflation",
    "oper_acq": "Operating acquisition ($)",
    "marketing_acq": "Marketing acquisition ($)",
    "maintenance": "Maintenance ($)",
    "premium_tax": "Premium tax",
    "tax_rate": "Tax rate",
    "ibnr_pct": "IBNR as % of claims",
    "rbc_pct_of_prem": "RBC as % of premium",
    "rbc_factor": "RBC factor",
    "covariance": "Covariance",
}


def _economic(rows) -> dict:
    return {field: float(_kv(rows, label)) for field, label in _ECONOMIC_LABELS.items()}


def assumptions_from_workbook(file) -> dict:
    """Parse a workbook (path or file-like) into the canonical assumptions dict."""
    from openpyxl import load_workbook

    wb = load_workbook(file, data_only=True)
    missing = [s for s in SHEET_NAMES if s not in wb.sheetnames and s != "Derived factors"]
    if missing:
        raise ValueError(f"Workbook is missing expected sheet(s): {', '.join(missing)}")

    def rows(name):
        return list(wb[name].iter_rows(values_only=True))

    return {
        "schema_version": str(_kv(rows("Overview"), "Schema version") or "1"),
        "pull_forward": _pull_forward(rows("Pull forward")),
        "morbidity": _morbidity(rows("Morbidity")),
        "premium": _premium(rows("Premium")),
        "rerates": _rerates(rows("Rerates")),
        "distribution": _distribution(rows("Distribution")),
        "termination": _termination(rows("Termination")),
        "commission": _commission(rows("Commission")),
        "other": _economic(rows("Economic")),
    }
