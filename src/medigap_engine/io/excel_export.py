"""Export an :class:`AssumptionSet` to a multi-sheet Excel workbook.

One sheet per assumption category (mirroring the Assumptions tab sub-tabs) plus a
**Derived factors** sheet showing the values the engine actually applies — the
raw inputs are differentials/relativities, but the engine first converts them to
mix-normalised factors (see ``normalized_factors`` / ``derive_two_level`` in
``models.assumptions``). Exporting both lets the model be rebuilt and checked in
Excel.

``openpyxl`` is imported lazily inside the functions so this module (and the rest
of the engine) stays importable with numpy only, preserving the Pyodide/engine
purity constraint; ``openpyxl`` is only needed when an export is actually run.
"""
from __future__ import annotations

import datetime as _dt
import io

from ..models.assumptions import (
    PROJECTION_YEARS,
    AssumptionSet,
    derive_two_level,
    normalized_factors,
)

# sheet names in workbook order (also the Overview index)
SHEET_NAMES = [
    "Overview",
    "Pull forward",
    "Morbidity",
    "Premium",
    "Rerates",
    "Distribution",
    "Termination",
    "Commission",
    "Economic",
    "Derived factors",
]


# --- small sheet-writing helpers --------------------------------------------
# Each takes the worksheet and the next free row, writes a block, and returns the
# next free row, so sheet builders read top-to-bottom.

def _title(ws, row: int, text: str):
    from openpyxl.styles import Font

    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12)
    return row + 1


def _header(ws, row: int, labels):
    from openpyxl.styles import Font

    for j, label in enumerate(labels, start=1):
        ws.cell(row=row, column=j, value=label).font = Font(bold=True)
    return row + 1


def _kv(ws, row: int, title: str, items):
    """Write a titled block of (label, value) pairs in two columns."""
    row = _title(ws, row, title)
    for label, value in items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1  # trailing blank row


def _table(ws, row: int, title: str, headers, data_rows):
    """Write a titled table: bold header row then ``data_rows`` (each a sequence)."""
    row = _title(ws, row, title)
    row = _header(ws, row, headers)
    for r in data_rows:
        for j, value in enumerate(r, start=1):
            ws.cell(row=row, column=j, value=value)
        row += 1
    return row + 1


def _durations(n: int):
    return list(range(1, n + 1))


# --- per-category sheet builders --------------------------------------------

def _overview(ws, a: AssumptionSet) -> None:
    row = _title(ws, 1, "Medigap pricing model — assumptions export")
    row += 1
    row = _kv(ws, row, "About", [
        ("Schema version", a.schema_version),
        ("Projection years", PROJECTION_YEARS),
        ("Exported (UTC date)", _dt.datetime.utcnow().strftime("%Y-%m-%d")),
    ])
    _table(ws, row, "Sheets in this workbook",
           ["#", "Sheet"],
           [(i, name) for i, name in enumerate(SHEET_NAMES, start=1)])


def _pull_forward(ws, a: AssumptionSet) -> None:
    pf = a.pull_forward
    _kv(ws, 1, "Pull experience forward to the pricing period", [
        ("Duration (years to pull forward)", pf.duration),
        ("Claims trend (pull-forward)", pf.claims_trend),
        ("Premium trend (pull-forward)", pf.premium_trend),
    ])


def _morbidity(ws, a: AssumptionSet) -> None:
    m = a.morbidity
    row = _table(
        ws, 1, "Base claim cost by attained age and plan (gender blend)",
        ["Age", *m.plans],
        [[m.ages[i], *[m.base_cc[p][i] for p in m.plans]] for i in range(len(m.ages))],
    )
    row = _kv(ws, row, "Claim-cost differentials", [
        ("Male claim cost higher than female by", m.gender_cc_diff),
        ("Non-preferred higher than preferred by", m.preferred_diff),
        ("Non-HHD higher than HHD by", m.hhd_diff),
    ])
    row = _table(ws, row, "Projection trend & claim-cost aging by duration",
                 ["Duration", "Trend", "CC aging"],
                 [[d, m.trend_by_year[i], m.cc_aging_by_duration[i]]
                  for i, d in enumerate(_durations(len(m.trend_by_year)))])
    row = _table(ws, row, "State morbidity factors", ["State", "Factor"],
                 sorted(m.state_factors.items()))
    if m.preferred_factors or m.hhd_factors:
        levels = sorted(set(m.preferred_factors) | set(m.hhd_factors))
        row = _table(ws, row, "Raw preferred / HHD claim factors by level",
                     ["Level", "Preferred factor", "HHD factor"],
                     [[lvl, m.preferred_factors.get(lvl), m.hhd_factors.get(lvl)]
                      for lvl in levels])
    if m.selection_factors:
        keys = ["duration", "issue_age", "uw", "factor"]
        _table(ws, row, "UW selection factors", keys,
               [[r.get(k) for k in keys] for r in m.selection_factors])


def _premium(ws, a: AssumptionSet) -> None:
    p = a.premium
    row = _table(ws, 1, "Base premium by issue age (plan-G blend)",
                 ["Issue age", "Base premium"], sorted(p.base_by_issue_age.items()))
    row = _table(ws, row, "Plan relativities (G = 1.00)", ["Plan", "Relativity"],
                 list(p.plan_rel.items()))
    row = _table(ws, row, "UW relativities", ["UW class", "Relativity"],
                 list(p.uw_rel.items()))
    row = _kv(ws, row, "Premium differentials", [
        ("Male premium higher than female by", p.gender_diff),
        ("Non-preferred premium higher by", p.preferred_diff),
        ("Non-HHD premium higher by", p.hhd_diff),
    ])
    row = _table(ws, row, "State premium factors (raw)", ["State", "Factor"],
                 sorted(p.state_factor.items()))
    if p.cell_premiums:
        states = sorted({s for m in p.cell_premiums.values() for s in m})
        rows = [[label, *[m[s] if s in m else None for s in states]]   # full precision: exact
                for label, m in sorted(p.cell_premiums.items())]
        _table(ws, row, "Per-cell premiums — exact rates by state (override the factor model)",
               ["Cell", *states], rows)


def _rerates(ws, a: AssumptionSet) -> None:
    r = a.rerates
    row = _kv(ws, 1, "Rerate strategy & rules", [
        ("Solve to target lifetime LR", r.solve),
        ("Target lifetime LR", r.target_lifetime_lr),
        ("Target IRR (reported)", r.target_irr),
        ("Antiselection lambda — claims", r.antiselection_lambda_claims),
        ("Antiselection lambda — lapse", r.antiselection_lambda_lapse),
        ("Max single rerate", r.max_rerate),
        ("In-year LR floor", r.in_year_lr_floor),
        ("Consecutive rule: z", r.consecutive_z),
        ("Consecutive rule: b (years)", r.consecutive_b),
    ])
    row = _table(ws, row, "Specified rerates by duration (shared / national)",
                 ["Duration", "Rerate"],
                 [[d, r.specified_rerates[i]]
                  for i, d in enumerate(_durations(len(r.specified_rerates)))])
    row = _table(ws, row, "Aging rerate by attained age (column H)", ["Age", "Aging rerate"],
                 list(zip(r.aging_rerate_by_age_ages, r.aging_rerate_by_age_factor)))
    if r.by_state:
        states = sorted(r.by_state)
        rows = [[d, *[r.by_state[s][i] if i < len(r.by_state[s]) else None for s in states]]
                for i, d in enumerate(_durations(PROJECTION_YEARS))]
        row = _table(ws, row, "Per-state rerate overrides by duration",
                     ["Duration", *states], rows)
    if r.target_lifetime_lr_by_state:
        _table(ws, row, "Per-state target lifetime LR overrides", ["State", "Target LR"],
               sorted(r.target_lifetime_lr_by_state.items()))


def _distribution(ws, a: AssumptionSet) -> None:
    d = a.distribution
    uws = list(d.uw)
    ages = sorted(d.by_issue_age)
    row = _title(ws, 1, "Joint distribution weight grid (plan x issue age x UW); "
                        "gender / preferred / HHD are marginals applied on top")
    row += 1
    # one grid table per plan: rows = issue age, columns = UW class
    for pl, grid in d.joint.items():
        data = [[age, *[grid.get(str(age), {}).get(u, 0.0) for u in uws]] for age in ages]
        row = _table(ws, row, f"Plan {pl} weight grid (issue age x UW)",
                     ["Issue age", *uws], data)
    for title, mapping in (("Gender", d.gender), ("Preferred", d.preferred), ("HHD", d.hhd)):
        items = sorted(mapping.items())
        rows = [[k, v] for k, v in items]
        rows.append(["Sum", sum(mapping.values())])
        row = _table(ws, row, title, ["Level", "Weight"], rows)
    if d.sep_rule_states:
        row = _table(ws, row, "Special Enrollment Period (SEP) states",
                     ["State"], [[s] for s in sorted(d.sep_rule_states)])
    if d.state_weights:
        row = _table(ws, row, "New-business volume weights by state (combine weighting)",
                     ["State", "Weight"], sorted(d.state_weights.items()))
    if d.by_state:
        rows = []
        for s in sorted(d.by_state):
            joint = d.by_state[s].get("joint", {})
            for pl in sorted(joint):
                for age in ages:
                    cell = joint.get(pl, {}).get(str(age), {})
                    rows.append([s, pl, age, *[cell.get(u, 0.0) for u in uws]])  # exact weights
        if rows:
            row = _table(ws, row, "Per-state mix-of-business grid (joint plan x issue age x UW)",
                         ["State", "Plan", "Issue age", *uws], rows)
        # per-state gender / preferred / HHD marginals (grid_weight multiplies these in)
        marg = []
        for s in sorted(d.by_state):
            for dim in ("gender", "preferred", "hhd"):
                for lvl, w in sorted(d.by_state[s].get(dim, {}).items()):
                    marg.append([s, dim, str(lvl), float(w)])   # exact weights
        if marg:
            _table(ws, row, "Per-state gender / preferred / HHD marginals",
                   ["State", "Dimension", "Level", "Weight"], marg)


def _termination(ws, a: AssumptionSet) -> None:
    t = a.termination
    row = _table(ws, 1, "Base lapse (blend) and UW relativity by duration",
                 ["Duration", "Base lapse (blend)", "UW relativity"],
                 [[d, t.base_lapse[i], t.uw_lapse_rel[i]]
                  for i, d in enumerate(_durations(len(t.base_lapse)))])
    row = _kv(ws, row, "Termination duration scaling", [
        ("Duration 2 scaling", t.dur2_scaling),
        ("Duration 3+ scaling", t.dur3plus_scaling),
    ])
    row = _table(ws, row, "State lapse factors", ["State", "Factor"],
                 sorted(t.state_factors.items()))
    _table(ws, row, "Mortality table", ["Age", "qx"], list(zip(t.mort_age, t.mort_qx)))


def _commission(ws, a: AssumptionSet) -> None:
    c = a.commission
    states = list(c.by_state.keys())
    rows = []
    for i, d in enumerate(_durations(PROJECTION_YEARS)):
        rows.append([d, *[c.by_state[s][i] if i < len(c.by_state[s]) else None
                          for s in states]])
    row = _table(ws, 1, "Commission rate by state and duration",
                 ["Duration", *states], rows)
    row = _table(ws, row, "National schedules by duration",
                 ["Duration", "Plan N", "Non-N"],
                 [[d, c.plan_n_schedule[i] if i < len(c.plan_n_schedule) else None,
                   c.nonn_schedule[i] if i < len(c.nonn_schedule) else None]
                  for i, d in enumerate(_durations(PROJECTION_YEARS))])
    _kv(ws, row, "Commission rules", [
        ("GI flat commission", c.gi_flat),
        ("Plan F premium offset", c.plan_f_offset),
        ("Halve commission for issue age >= 80", c.age80_halving),
    ])


def _economic(ws, a: AssumptionSet) -> None:
    o = a.other
    _kv(ws, 1, "Economic assumptions", [
        ("Discount rate", o.discount_rate),
        ("NIER (investment return)", o.nier),
        ("Inflation", o.inflation),
        ("Operating acquisition ($)", o.oper_acq),
        ("Marketing acquisition ($)", o.marketing_acq),
        ("Maintenance ($)", o.maintenance),
        ("Premium tax", o.premium_tax),
        ("Tax rate", o.tax_rate),
        ("IBNR as % of claims", o.ibnr_pct),
        ("RBC as % of premium", o.rbc_pct_of_prem),
        ("RBC factor", o.rbc_factor),
        ("Covariance", o.covariance),
    ])


def _derived(ws, a: AssumptionSet) -> None:
    """The mix-normalised factors the engine actually applies. Mirrors the
    captions on the Assumptions tab so the two agree exactly."""
    m, p, d, t, pf = (
        a.morbidity, a.premium, a.distribution, a.termination, a.pull_forward,
    )
    row = _title(ws, 1, "Derived factors actually used by the engine")
    row += 1
    row = _kv(ws, row, "Pull-forward bring-forward factors  (1 + trend) ^ duration", [
        ("Claims bring-forward", (1.0 + pf.claims_trend) ** pf.duration),
        ("Premium bring-forward", (1.0 + pf.premium_trend) ** pf.duration),
    ])

    pg = normalized_factors({"M": 1.0 + p.gender_diff, "F": 1.0}, d.gender)
    pp = normalized_factors({"N": 1.0 + p.preferred_diff, "Y": 1.0}, d.preferred)
    ph = normalized_factors({"N": 1.0 + p.hhd_diff, "Y": 1.0}, d.hhd)
    puw = normalized_factors(p.uw_rel, d.uw)
    row = _table(
        ws, row, "Premium normalised factors (mix-weighted mean = 1)",
        ["Factor", "Level", "Value"],
        [["Gender", "M", pg["M"]], ["Gender", "F", pg["F"]],
         ["Preferred", "Y", pp["Y"]], ["Preferred", "N", pp["N"]],
         ["HHD", "Y", ph["Y"]], ["HHD", "N", ph["N"]],
         *[["UW", k, v] for k, v in puw.items()]],
    )

    mg = normalized_factors({"M": 1.0 + m.gender_cc_diff, "F": 1.0}, d.gender)
    mp = derive_two_level(d.preferred.get("Y", 0.5), m.preferred_diff)
    mh = derive_two_level(d.hhd.get("Y", 0.5), m.hhd_diff)
    row = _table(
        ws, row, "Morbidity (claim-cost) normalised factors",
        ["Factor", "Level", "Value"],
        [["Gender", "M", mg["M"]], ["Gender", "F", mg["F"]],
         ["Preferred", "Y", mp["Y"]], ["Preferred", "N", mp["N"]],
         ["HHD", "Y", mh["Y"]], ["HHD", "N", mh["N"]]],
    )

    tuw = normalized_factors(
        {"UW": t.uw_lapse_rel[0], "OE": 1.0, "GI": 1.0}, d.uw)
    _table(ws, row, "Termination duration-1 applied UW lapse factors",
           ["UW class", "Value"], list(tuw.items()))


_BUILDERS = {
    "Overview": _overview,
    "Pull forward": _pull_forward,
    "Morbidity": _morbidity,
    "Premium": _premium,
    "Rerates": _rerates,
    "Distribution": _distribution,
    "Termination": _termination,
    "Commission": _commission,
    "Economic": _economic,
    "Derived factors": _derived,
}


def assumptions_to_workbook(a: AssumptionSet):
    """Build an ``openpyxl`` Workbook with one sheet per assumption category plus a
    Derived-factors sheet. ``openpyxl`` is imported here (lazily) on purpose."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    for name in SHEET_NAMES:
        ws = wb.create_sheet(title=name)
        _BUILDERS[name](ws, a)
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 16
    return wb


def assumptions_to_xlsx_bytes(a: AssumptionSet) -> bytes:
    """Serialise the assumptions workbook to ``.xlsx`` bytes (for a download button)."""
    buf = io.BytesIO()
    assumptions_to_workbook(a).save(buf)
    return buf.getvalue()
