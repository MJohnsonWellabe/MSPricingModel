"""Extract clean sales/claims template + sample CSVs from the source workbook.

Usage:
    python tools/extract_samples.py path/to/MS_Pricing_By_State_2026AEP_v5.xlsm.xlsx [claims_cap]

Writes (under src/medigap_engine/data/templates/):
    sales_template.csv, sales_sample.csv, claims_template.csv, claims_sample.csv

The SalesData tab's third pivot block carries Application Count + Entered Premium
alongside already-bucketed cell dimensions; ClaimsData columns A:Q are the raw
claim rows.
"""
from __future__ import annotations

import csv
import os
import sys

import openpyxl

HERE = os.path.dirname(__file__)
OUT = os.path.join(HERE, "..", "src", "medigap_engine", "data", "templates")

SALES_HEADER = ["state", "issue_age", "gender", "plan", "uw_class",
                "preferred", "hhd", "application_count", "entered_premium"]
CLAIMS_HEADER = ["state", "plan", "issue_age", "gender", "uw_class",
                 "duration", "cnt", "earned", "annualized_prem", "adj_claims"]


def _f(v):
    try:
        s = str(v).strip()
        if s == "" or s.upper() == "NULL":
            return 0.0
        return float(s)
    except (TypeError, ValueError):
        return 0.0


def _write(path, header, rows):
    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)


def extract_sales(ws):
    # third block: 0-based U=20, Y=24, Z=25, AA=26..AF=31
    rows = []
    for r in ws.iter_rows(min_row=9, values_only=True):
        if len(r) <= 31 or r[26] is None:  # AA = issue age (bucketed)
            continue
        state, age, gender, plan, uw, pref, hhd = (
            r[20], r[26], r[27], r[28], r[29], r[30], r[31])
        count, prem = r[24], r[25]
        if plan is None or gender is None:
            continue
        rows.append([state, age, gender, plan, uw, pref, hhd,
                     count or 0, round(float(prem or 0), 2)])
    return rows


def extract_claims(ws, cap):
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None or r[12] is None:  # earned / Plan letter
            continue
        earned, claims_, _date, _pc, state, _idate, age, sex, ann, cnt, uw, \
            _inc, plan, _bucket, adj, dur, _sep = r[:17]
        rows.append([state, plan, age, sex, uw, dur, cnt,
                     round(_f(earned), 2), round(_f(ann), 2), round(_f(adj), 2)])
        if len(rows) >= cap:
            break
    return rows


def main(path, cap=25000):
    os.makedirs(OUT, exist_ok=True)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sales = extract_sales(wb["SalesData"])
    claims = extract_claims(wb["ClaimsData"], cap)
    wb.close()
    _write(os.path.join(OUT, "sales_sample.csv"), SALES_HEADER, sales)
    _write(os.path.join(OUT, "sales_template.csv"), SALES_HEADER, sales[:8])
    _write(os.path.join(OUT, "claims_sample.csv"), CLAIMS_HEADER, claims)
    _write(os.path.join(OUT, "claims_template.csv"), CLAIMS_HEADER, claims[:8])
    print(f"sales rows: {len(sales)}; claims rows: {len(claims)} (cap {cap})")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    cap = int(sys.argv[2]) if len(sys.argv) > 2 else 25000
    main(sys.argv[1], cap)
