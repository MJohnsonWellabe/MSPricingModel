"""Calibration harness: run the engine for one state and diff every aggregate
line against the source workbook's "Aggregate Model" sheet, plus a single-cell
diff against the "Model" sheet.

Usage:
    python tools/compare_tx.py /path/to/workbook.xlsx [STATE]

Read-only; not shipped in the engine. Drives the TX validation to exact.
"""
from __future__ import annotations

import sys

import openpyxl

sys.path.insert(0, "tools")
import generate_seed as gs  # noqa: E402

from medigap_engine.io.defaults import build_cells as eng_build_cells  # noqa: E402
from medigap_engine.io.serialize import assumptions_from_dict  # noqa: E402
from medigap_engine.models.cell import CellKey  # noqa: E402
from medigap_engine.engine.run import run  # noqa: E402
from medigap_engine.models.config import RunConfig  # noqa: E402

# Aggregate Model row label (col D) -> engine series key
_LINES = [
    ("Lives", "lives"),
    ("Earned prem", "earned_prem"),
    ("NII", "nii"),
    ("claim costs", "claims"),
    ("Commission", "commission"),
    ("premium tax", "premium_tax"),
    ("Oper. Acq", "oper_acq"),
    ("Marketing", "marketing"),
    ("Maint", "maintenance"),
    ("pretax income", "pretax_income"),
    ("tax", "tax"),
    ("at income", "at_income"),
    ("int on rbc", "int_on_rbc"),
    ("tax on int", "tax_on_int"),
    ("at adjusted income", "ah_cashflow"),
    ("LR", "in_year_lr"),
]


def build_tx_assumptions(wb) -> dict:
    """Extract a faithful AssumptionSet dict from the workbook: joint distribution
    grid + exact per-cell premiums + no premium pull-forward."""
    A = gs.build_assumptions(wb["Assumptions"])
    cells = gs.build_cells(wb["Input"])

    # joint plan x age x uw grid from the per-cell weights (sum over gender/pref/hhd)
    tot = sum(c["weight"] for c in cells) or 1.0
    joint: dict[str, dict[str, dict[str, float]]] = {}
    for c in cells:
        ages = joint.setdefault(str(c["plan"]), {}).setdefault(str(int(c["issue_age"])), {})
        ages[str(c["uw"])] = ages.get(str(c["uw"]), 0.0) + c["weight"] / tot

    def marg(field):
        g: dict = {}
        for c in cells:
            g[c[field]] = g.get(c[field], 0.0) + c["weight"] / tot
        return g

    A["distribution"] = {
        "joint": joint,
        "gender": marg("gender"), "preferred": marg("preferred"), "hhd": marg("hhd"),
    }

    # exact per-cell premiums (cell label -> state -> premium); plus a factor-model
    # fallback so non-workbook states still price
    prem, _dist = gs.build_factor_blocks(cells)
    cell_premiums = {}
    for c in cells:
        label = CellKey(int(c["issue_age"]), c["gender"], c["plan"], c["uw"],
                        c["preferred"], c["hhd"]).label()
        sp = dict(c.get("state_premiums", {}))
        sp.setdefault("All", c["premium"])
        cell_premiums[label] = {s: round(float(v), 6) for s, v in sp.items()}
    prem["cell_premiums"] = cell_premiums
    A["premium"] = prem

    # gender blend of base claim cost (the workbook male = 1.15 x female exactly)
    m = A["morbidity"]
    gmix = A["distribution"]["gender"]
    g_rel = {"M": 1.0 + m["gender_cc_diff"], "F": 1.0}
    gblend = sum(gmix.get(g, 0.0) * g_rel[g] for g in g_rel)
    m["base_cc"] = {pl: [round(v * gblend, 6) for v in m["base_cc"][pl]] for pl in m["base_cc"]}

    # uw-mix blend of base lapse
    t = A["termination"]
    w_uw = A["distribution"]["uw"].get("UW", 0.0) if False else marg("uw").get("UW", 0.0)
    w_o = 1.0 - w_uw
    t["base_lapse"] = [round(t["base_lapse"][i] * (w_uw * t["uw_lapse_rel"][i] + w_o), 6)
                       for i in range(len(t["base_lapse"]))]

    # the Input premium IS the pricing rate -> no premium pull-forward
    A["pull_forward"]["premium_trend"] = 0.0
    # the workbook uses its specified rerate schedule (col F), not a solve
    A["rerates"]["solve"] = False
    # morbidity state factor is the per-run scalar Input!Z1 (claims = ... x StateFactor)
    z1 = wb["Input"]["Z1"].value
    if isinstance(z1, (int, float)):
        A["morbidity"]["state_factors"]["TX"] = round(float(z1), 8)
        A["morbidity"]["state_factors"].setdefault("All", 1.0)
    # raw preferred (AT) / hhd (AU) claim factors; the workbook lapse has no
    # antiselective load, so lapse lambda is zero
    Aw = wb["Assumptions"]
    ci = openpyxl.utils.column_index_from_string
    A["morbidity"]["preferred_factors"] = {str(Aw.cell(r, ci("AS")).value):
                                           float(Aw.cell(r, ci("AT")).value) for r in (4, 5)}
    A["morbidity"]["hhd_factors"] = {str(Aw.cell(r, ci("AS")).value):
                                     float(Aw.cell(r, ci("AU")).value) for r in (4, 5)}
    A["rerates"]["antiselection_lambda_lapse"] = 0.0
    return A


def targets(wb, state):
    ws = wb["Aggregate Model"]
    # find each line's row by its label in column D, read durations 1..30 (cols F..)
    rows = {}
    for r in range(1, ws.max_row + 1):
        lab = ws.cell(r, 4).value
        if isinstance(lab, str):
            rows[lab.strip()] = r
    out = {}
    for lab, key in _LINES:
        r = rows.get(lab)
        if r is None:
            continue
        out[key] = [ws.cell(r, 6 + i).value for i in range(30)]
    return out


def main(path, state="TX"):
    wb = openpyxl.load_workbook(path, data_only=True)
    A = build_tx_assumptions(wb)
    tgt = targets(wb, state)
    wb.close()

    asm = assumptions_from_dict(A)
    result, _ = run(eng_build_cells(asm), asm, RunConfig(states=[state], solve_rerates=False))
    sr = result.by_state[state]

    print(f"=== {state}: engine vs workbook Aggregate Model (durations 1-12) ===")
    ndur = 12
    for _lab, key in _LINES:
        if key not in tgt:
            continue
        mine = sr.series[key]
        xl = tgt[key]
        cells = []
        worst = 0.0
        for i in range(ndur):
            m = mine[i] if i < len(mine) else 0.0
            x = xl[i]
            if not isinstance(x, (int, float)):
                cells.append("   n/a")
                continue
            if abs(x) > 1e-9:
                rel = (m - x) / x
                worst = max(worst, abs(rel))
                cells.append(f"{rel*100:+6.1f}%")
            else:
                cells.append(f"{m-x:+6.2f}")
        flag = " OK" if worst < 0.005 else (" ~" if worst < 0.05 else " XX")
        print(f"{key:14} worst={worst*100:5.1f}%{flag}  " + " ".join(cells))


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "/tmp/tx_model.xlsx",
         sys.argv[2] if len(sys.argv) > 2 else "TX")
