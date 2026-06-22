"""Regenerate the bundled seed data from the source pricing workbook.

Usage:
    python tools/generate_seed.py path/to/MS_Pricing_By_State_2026AEP_v5.xlsm.xlsx

Writes:
    src/medigap_engine/data/default_assumptions.json
    src/medigap_engine/data/default_cells.json

Requires openpyxl (``pip install openpyxl``). This script encodes how each
assumption block maps out of the workbook; see the README for the mapping.
"""
from __future__ import annotations

import json
import os
import sys

import openpyxl

HERE = os.path.dirname(__file__)
DATA = os.path.join(HERE, "..", "src", "medigap_engine", "data")
PLANS = ["F", "G", "N"]

# per-state premium columns on the Input tab (state -> column letter)
STATE_COLS = {
    "TX": "T", "All": "U", "WI": "V", "DE": "W", "KY": "X", "NH": "Y", "OH": "Z",
    "LA": "AA", "KS": "AB", "FL": "AC", "TN": "AD", "NE": "AE", "MD": "AF",
    "MO": "AG", "CO": "AH", "GA": "AI", "CA": "AJ", "AZ": "AK", "PA": "AL",
    "MI": "AM", "NC": "AN", "WA": "AO", "NJ": "AP", "IN": "AQ", "VA": "AR",
    "IA": "AS", "SC": "AT", "IL": "AV",
}


def _rnd(v, n=6):
    try:
        return round(float(v), n)
    except (TypeError, ValueError):
        return None


def build_assumptions(A) -> dict:
    ci = openpyxl.utils.column_index_from_string

    def c(letter, r):
        return A.cell(r, ci(letter)).value

    ages = [int(c("Q", r)) for r in range(14, 50)]

    def cc(cols):
        m = {"F": cols[0], "G": cols[1], "N": cols[2]}
        return {p: [_rnd(c(m[p], r), 4) for r in range(14, 50)] for p in PLANS}

    selection = []
    for r in range(4, 95):
        yr = c("AN", r)
        if yr is None:
            continue
        selection.append({
            "duration": int(yr), "issue_age": int(c("AO", r)),
            "uw": c("AP", r), "factor": _rnd(c("AQ", r), 6),
        })

    state_factors = {}
    for r in range(4, 33):
        s = c("AZ", r)
        if s is not None:
            state_factors[str(s)] = _rnd(c("BA", r), 4)

    comm_by_state = {}
    for col in range(ci("BX"), ci("CY") + 1):
        s = A.cell(2, col).value
        if s is None:
            continue
        comm_by_state[str(s)] = [_rnd(A.cell(r, col).value, 4) for r in range(3, 33)]

    sc = {A.cell(r, 1).value: A.cell(r, 2).value for r in range(2, 14)}

    # gender claim cost: male = 1.15 x female; store female as the base table
    cc_female = cc(["AE", "AF", "AG"])
    cc_male = cc(["Y", "Z", "AA"])
    gender_ratio = round(cc_male["G"][0] / cc_female["G"][0], 6) if cc_female["G"][0] else 1.15
    # preferred / hhd differentials (no-level over yes-level), from the workbook factors
    pref_y, pref_n = _rnd(c("AT", 4)), _rnd(c("AT", 5))
    hhd_y, hhd_n = _rnd(c("AU", 4)), _rnd(c("AU", 5))
    # OE/GI ("other") lapse and the UW factor by duration
    oe_lapse = [_rnd(c("BM", r)) for r in range(3, 33)]
    uw_lapse = [_rnd(c("BO", r)) for r in range(3, 33)]
    uw_factor = [round(uw_lapse[i] / oe_lapse[i], 6) if oe_lapse[i] else 1.0
                 for i in range(len(oe_lapse))]

    trend_by_year = [_rnd(c("G", r), 4) for r in range(3, 33)]
    # raw preferred (AT) / hhd (AU) claim factors keyed by class level (AS)
    pref_factors = {str(c("AS", r)): _rnd(c("AT", r), 8) for r in (4, 5)}
    hhd_factors = {str(c("AS", r)): _rnd(c("AU", r), 8) for r in (4, 5)}
    return {
        "schema_version": "2",
        "pull_forward": {
            "duration": 1.75,
            "claims_trend": trend_by_year[0],
            # the Input premium is already the pricing rate -> no premium pull-forward
            "premium_trend": 0.0,
        },
        "morbidity": {
            "ages": ages, "plans": PLANS,
            # base_cc is converted to the gender blend in main() using the gender mix
            "base_cc": cc_female,
            "gender_cc_diff": round(gender_ratio - 1.0, 6),
            "state_factors": state_factors,
            "selection_factors": selection,
            "cc_aging_by_duration": [_rnd(c("AL", r), 6) or 0.0 for r in range(3, 33)],
            "preferred_diff": round(pref_n / pref_y - 1, 6),
            "hhd_diff": round(hhd_n / hhd_y - 1, 6),
            "trend_by_year": trend_by_year,
            "preferred_factors": pref_factors,
            "hhd_factors": hhd_factors,
        },
        "rerates": {
            # default to solving so every state prices sensibly; the TX validation
            # turns solve OFF to use the workbook's specified rerate schedule (col F)
            "solve": True,
            "specified_rerates": [_rnd(c("F", r), 4) for r in range(3, 33)],
            "aging_rerate_by_age_ages": [int(c("AI", r)) for r in range(3, 39)],
            "aging_rerate_by_age_factor": [_rnd(c("AJ", r), 6) for r in range(3, 39)],
            "target_lifetime_lr": 0.78, "target_irr": 0.15,
            "max_rerate": 0.20, "in_year_lr_floor": 0.65,
            "consecutive_z": 0.15, "consecutive_b": 5,
            # claims antiselection P = (1+aging)P + 0.5*(rerate-trend); the workbook
            # lapse has no antiselective load, so the lapse lambda is zero
            "antiselection_lambda_claims": 0.5, "antiselection_lambda_lapse": 0.0,
        },
        # "premium" and "distribution" are factor models derived from the cell
        # universe; see build_factor_blocks() and main().
        "termination": {
            # base_lapse is converted to the uw-mix blend in main()
            "base_lapse": oe_lapse,
            "uw_lapse_rel": uw_factor,
            "state_factors": {k: 1.0 for k in state_factors},
            "mort_age": [int(c("BS", r)) for r in range(3, 104)],
            "mort_qx": [_rnd(c("BT", r), 8) for r in range(3, 104)],
            "dur2_scaling": 1.05, "dur3plus_scaling": 1.10,
        },
        "commission": {
            "by_state": comm_by_state,
            "plan_n_schedule": [_rnd(c("E", r), 4) for r in range(33, 63)],
            "nonn_schedule": [_rnd(c("E", r), 4) for r in range(3, 33)],
            "gi_flat": 25.0, "plan_f_offset": 240.0, "age80_halving": True,
        },
        "other": {
            "discount_rate": _rnd(sc.get("Discount"), 4),
            "premium_tax": _rnd(sc.get("premium tax"), 4),
            "oper_acq": sc.get("Oper. Acq"), "marketing_acq": sc.get("Marketing"),
            "maintenance": sc.get("Maint"), "inflation": _rnd(sc.get("inflation"), 4),
            "rbc_factor": sc.get("RBC"), "covariance": _rnd(sc.get("Covariance"), 4),
            "rbc_pct_of_prem": _rnd(sc.get("RBC as % of prem"), 4),
            "nier": _rnd(sc.get("NIER"), 4), "tax_rate": _rnd(sc.get("Tax rate"), 4),
            "ibnr_pct": _rnd(sc.get("IBNR % of Claims"), 4),
        },
    }


def build_cells(ws) -> list:
    ci = openpyxl.utils.column_index_from_string
    state_idx = {s: ci(col) - 1 for s, col in STATE_COLS.items()}
    cells = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            break
        sp = {s: round(float(row[i]), 4) for s, i in state_idx.items()
              if i < len(row) and isinstance(row[i], (int, float))}
        cells.append({
            "cell": int(row[0]), "issue_age": int(row[1]), "gender": row[2],
            "plan": row[3], "uw": row[4], "preferred": row[5], "hhd": row[6],
            "weight": round(float(row[7] or 0), 8), "premium": round(float(row[8] or 0), 4),
            "state_premiums": sp,
        })
    return cells


def build_factor_blocks(cells: list) -> tuple[dict, dict]:
    """Derive the premium factor model and distribution weight factors from the
    per-cell universe via an additive log main-effects decomposition.

    premium(cell, state) = base_by_issue_age[age] x gender x plan x uw x pref x hhd x state
    weight(cell)         = product of per-dimension marginal weights
    """
    import math
    from collections import defaultdict

    logs = [math.log(c["premium"]) for c in cells if c["premium"] > 0]
    mu = sum(logs) / len(logs)

    def eff(field):
        g = defaultdict(list)
        for c in cells:
            if c["premium"] > 0:
                g[c[field]].append(math.log(c["premium"]))
        return {k: sum(v) / len(v) - mu for k, v in g.items()}

    # plan anchored at G: base reflects the plan-G level; plan_rel relative to G
    plan_eff = eff("plan")
    plan_g = plan_eff.get("G", 0.0)
    base_by_issue_age = {int(a): round(math.exp(mu + e + plan_g), 4)
                         for a, e in sorted(eff("issue_age").items())}
    plan_rel = {k: round(math.exp(e - plan_g), 6) for k, e in plan_eff.items()}

    def rel(field):  # centred relativity (normalised against the mix at runtime)
        return {k: round(math.exp(e), 6) for k, e in eff(field).items()}

    state_logs = defaultdict(list)
    for c in cells:
        comp = c["premium"]
        for s, p in c.get("state_premiums", {}).items():
            if comp > 0 and p > 0:
                state_logs[s].append(math.log(p / comp))
    state_factor = {s: round(math.exp(sum(v) / len(v)), 6) for s, v in state_logs.items()}
    state_factor.setdefault("All", 1.0)

    def diff(field, high, low):
        e = eff(field)
        return round(math.exp(e.get(high, 0.0) - e.get(low, 0.0)) - 1.0, 6)

    premium = {
        "base_by_issue_age": base_by_issue_age,
        "plan_rel": plan_rel, "uw_rel": rel("uw"),
        "gender_diff": diff("gender", "M", "F"),
        "preferred_diff": diff("preferred", "N", "Y"),
        "hhd_diff": diff("hhd", "N", "Y"),
        "state_factor": state_factor,
    }

    def marginal(field):
        g = defaultdict(float)
        tot = sum(c["weight"] for c in cells) or 1.0
        for c in cells:
            g[c[field]] += c["weight"]
        return {k: round(v / tot, 8) for k, v in g.items()}

    distribution = {
        "by_issue_age": {int(k): v for k, v in marginal("issue_age").items()},
        "gender": marginal("gender"), "plan": marginal("plan"), "uw": marginal("uw"),
        "preferred": marginal("preferred"), "hhd": marginal("hhd"),
    }
    return premium, distribution


def build_joint_distribution(cells: list) -> dict:
    """Joint plan x issue-age x UW weight grid from the per-cell weights (summing
    over gender/preferred/hhd), plus the gender/preferred/hhd marginals. Captures
    the non-separable plan/age/UW mix the workbook actually has."""
    from collections import defaultdict

    tot = sum(c["weight"] for c in cells) or 1.0
    joint: dict = {}
    for c in cells:
        ages = joint.setdefault(str(c["plan"]), {}).setdefault(str(int(c["issue_age"])), {})
        ages[str(c["uw"])] = round(ages.get(str(c["uw"]), 0.0) + c["weight"] / tot, 8)

    def marg(field):
        g = defaultdict(float)
        for c in cells:
            g[c[field]] += c["weight"] / tot
        return {k: round(v, 8) for k, v in g.items()}

    return {"joint": joint, "gender": marg("gender"),
            "preferred": marg("preferred"), "hhd": marg("hhd")}


def build_cell_premiums(cells: list) -> dict:
    """Exact per-cell premiums (cell label -> state -> premium) straight from Input."""
    out = {}
    for c in cells:
        label = (f"{c['issue_age']}{c['gender']}-{c['plan']}-{c['uw']}"
                 f"-P{c['preferred']}-H{c['hhd']}")
        sp = dict(c.get("state_premiums", {}))
        sp.setdefault("All", c["premium"])
        out[label] = {s: round(float(v), 6) for s, v in sp.items()}
    return out


# states that use a separate community-rating rule (different UW mix); editable later
SEP_RULE_STATES = ["IN", "VA", "MO", "WA", "CA", "MD", "KY", "DE"]


def merge_multistate_tables(assumptions: dict) -> None:
    """Overlay the workbook-derived per-state values onto the full multi-state reference
    so a TX-focused workbook does not delete the other states' assumptions. The workbook
    values (present in ``assumptions``) win; reference fills in the rest."""
    ref_path = os.path.join(HERE, "multistate_reference.json")
    if not os.path.exists(ref_path):
        return
    with open(ref_path) as f:
        ref = json.load(f)

    def merge(table: dict, base: dict) -> dict:
        out = dict(base)
        out.update({k: v for k, v in table.items() if k != "state"})  # drop junk key
        return out

    m = assumptions["morbidity"]
    m["state_factors"] = merge(m.get("state_factors", {}), ref["morbidity_state_factors"])
    t = assumptions["termination"]
    t["state_factors"] = merge(t.get("state_factors", {}), ref["termination_state_factors"])
    c = assumptions["commission"]
    c["by_state"] = merge(c.get("by_state", {}), ref["commission_by_state"])
    p = assumptions["premium"]
    p["state_factor"] = merge(p.get("state_factor", {}), ref["premium_state_factor"])
    # sep-rule classification (editable per-state input, not a hardcoded constant)
    assumptions["distribution"]["sep_rule_states"] = list(SEP_RULE_STATES)


def main(path: str) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    assumptions = build_assumptions(wb["Assumptions"])
    cells = build_cells(wb["Input"])
    # morbidity state factor is the per-run scalar Input!Z1 (claims x StateFactor)
    z1 = wb["Input"]["Z1"].value
    wb.close()

    premium, distribution = build_factor_blocks(cells)
    # exact per-cell premiums + true joint distribution grid override the factor model
    premium["cell_premiums"] = build_cell_premiums(cells)
    assumptions["premium"] = premium
    distribution = build_joint_distribution(cells)
    assumptions["distribution"] = distribution

    # restore the full multi-state per-state tables (this workbook is TX-focused and only
    # carries a few states); workbook-derived values overlay the reference baseline.
    merge_multistate_tables(assumptions)
    if isinstance(z1, (int, float)):
        assumptions["morbidity"]["state_factors"]["TX"] = round(float(z1), 8)
        assumptions["morbidity"]["state_factors"].setdefault("All", 1.0)

    # claims base cost is indexed by ISSUE age, so keep only the issue-age bands the
    # book actually prices (the distinct issue ages in the cell universe); the rest of
    # the 65-100 attained-age range was only needed when base cost aged by duration.
    m = assumptions["morbidity"]
    band_ages = sorted({int(c["issue_age"]) for c in cells})
    keep = [i for i, a in enumerate(m["ages"]) if a in band_ages]
    if keep:
        m["ages"] = [m["ages"][i] for i in keep]
        m["base_cc"] = {pl: [vals[i] for i in keep] for pl, vals in m["base_cc"].items()}

    # convert morbidity base_cc (female) -> gender blend, and termination base_lapse
    # (OE) -> uw-mix blend, now that the distribution mix is known
    gmix = distribution["gender"]
    g_rel = {"M": 1.0 + m["gender_cc_diff"], "F": 1.0}
    gblend = sum(gmix.get(g, 0.0) * g_rel[g] for g in g_rel)
    m["base_cc"] = {pl: [round(v * gblend, 4) for v in m["base_cc"][pl]] for pl in m["base_cc"]}
    t = assumptions["termination"]
    # uw marginal from the joint grid (sum over plan, age)
    uw_marg: dict = {}
    for ages in distribution["joint"].values():
        for uws in ages.values():
            for u, w in uws.items():
                uw_marg[u] = uw_marg.get(u, 0.0) + w
    w_uw = uw_marg.get("UW", 0.0)
    w_other = 1.0 - w_uw
    t["base_lapse"] = [round(t["base_lapse"][i] * (w_uw * t["uw_lapse_rel"][i] + w_other), 6)
                       for i in range(len(t["base_lapse"]))]

    # order keys for readability
    ordered = {}
    for k in ("schema_version", "pull_forward", "morbidity", "premium", "rerates",
              "distribution", "termination", "commission", "other"):
        if k in assumptions:
            ordered[k] = assumptions[k]
    for k in assumptions:
        ordered.setdefault(k, assumptions[k])

    with open(os.path.join(DATA, "default_assumptions.json"), "w") as fh:
        json.dump(ordered, fh, indent=1)
    with open(os.path.join(DATA, "default_cells.json"), "w") as fh:
        json.dump(cells, fh, indent=0)
    print(f"Wrote {len(cells)} cells and factor-based assumptions to {DATA}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1])
